import argparse
import json
import os
import random
import sys
import time
import hashlib
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import requests
from pydantic import BaseModel, ValidationError, field_validator
from urllib.parse import urljoin, urlparse
from playwright.sync_api import sync_playwright, Page, Browser, BrowserContext, TimeoutError as PlaywrightTimeoutError
from openpyxl import Workbook, load_workbook


# -----------------------------
# Data models
# -----------------------------

class Extraction(BaseModel):
    role_title: Optional[str] = None
    min_years_experience: Optional[int] = None
    max_years_experience: Optional[int] = None
    skills: List[str] = []
    location: Optional[str] = None
    job_type: Optional[str] = None  # full-time/part-time/intern/contract
    contact: Optional[str] = None   # email/URL if present
    verdict_relevant: bool

    @field_validator("skills", mode="before")
    @classmethod
    def ensure_list(cls, v: Any) -> List[str]:
        if v is None:
            return []
        if isinstance(v, list):
            return [str(x).strip() for x in v if str(x).strip()]
        if isinstance(v, str):
            # allow comma-separated string
            return [s.strip() for s in v.split(",") if s.strip()]
        return []


@dataclass
class PostRecord:
    timestamp_text: Optional[str]
    post_url: Optional[str]
    poster_name: Optional[str]
    poster_profile_url: Optional[str]
    poster_linkedin_id: Optional[str]
    role_title: Optional[str]
    min_years_experience: Optional[int]
    max_years_experience: Optional[int]
    skills: List[str]
    location: Optional[str]
    job_type: Optional[str]
    contact: Optional[str]
    post_excerpt: Optional[str]

    def to_row(self) -> Dict[str, Any]:
        return {
            "timestamp": self.timestamp_text or "",
            "post_url": self.post_url or "",
            "poster_name": self.poster_name or "",
            "poster_profile_url": self.poster_profile_url or "",
            "poster_linkedin_id": self.poster_linkedin_id or "",
            "role_title": self.role_title or "",
            "min_years_experience": self.min_years_experience,
            "max_years_experience": self.max_years_experience,
            "skills": ", ".join(self.skills) if self.skills else "",
            "location": self.location or "",
            "job_type": self.job_type or "",
            "contact": self.contact or "",
            "post_excerpt": self.post_excerpt or "",
        }


# -----------------------------
# Helpers
# -----------------------------

BASE_URL = "https://www.linkedin.com"


def sleep_random(min_seconds: float = 1.2, max_seconds: float = 3.0) -> None:
    time.sleep(random.uniform(min_seconds, max_seconds))


def normalize_whitespace(text: Optional[str]) -> Optional[str]:
    if text is None:
        return None
    return " ".join(text.split())


def ensure_full_url(href: Optional[str]) -> Optional[str]:
    if not href:
        return None
    if href.startswith("http://") or href.startswith("https://"):
        return href
    if href.startswith("/"):
        return urljoin(BASE_URL, href)
    return href


def extract_profile_id(profile_url: Optional[str]) -> Optional[str]:
    if not profile_url:
        return None
    try:
        parsed = urlparse(profile_url)
        # Typical profile path: /in/some-id/
        segments = [s for s in parsed.path.split("/") if s]
        if not segments:
            return None
        if segments[0] in {"in", "company"} and len(segments) >= 2:
            return segments[1]
        return segments[0]
    except Exception:
        return None


def text_hash(text: str) -> str:
    return hashlib.sha1(text.encode("utf-8", errors="ignore")).hexdigest()


# -----------------------------
# Ollama LLM client
# -----------------------------

class OllamaClient:
    def __init__(self, base_url: str = "http://localhost:11434", model: str = "llama3", request_timeout_sec: int = 60) -> None:
        self.base_url = base_url.rstrip("/")
        self.model = model
        self.request_timeout_sec = request_timeout_sec

    def build_prompt(self, post_text: str) -> str:
        return (
            "You extract hiring info from LinkedIn posts.\n"
            "Return strictly minified JSON only, no code fences or prose.\n"
            "Fields: role_title (string), min_years_experience (int), max_years_experience (int), "
            "skills (array of strings), location (string), job_type (full-time/part-time/intern/contract), "
            "contact (string), verdict_relevant (boolean: true only if role is Data Analyst or very close AND total experience required fits 0-2 years).\n"
            "If unsure about a field, use null, except skills should be [].\n"
            "Examples of relevant: 'Looking for a Data Analyst (freshers welcome)', 'Hiring Junior Data Analyst, 0-2 yrs'.\n"
            "Examples of NOT relevant: 'Senior Data Scientist 5+ years', 'Business Analyst 3-5 years'.\n\n"
            f"Text: \"\"\"{post_text}\"\"\"\n"
            "Respond with a single JSON object only."
        )

    def extract(self, post_text: str) -> Optional[Extraction]:
        # Try up to 2 attempts to get valid JSON
        for attempt in range(2):
            try:
                payload = {
                    "model": self.model,
                    "prompt": self.build_prompt(post_text),
                    "options": {"temperature": 0.1},
                    "stream": False,
                }
                resp = requests.post(
                    f"{self.base_url}/api/generate",
                    json=payload,
                    timeout=self.request_timeout_sec,
                )
                resp.raise_for_status()
                data = resp.json()
                raw = data.get("response", "").strip()
                # Strip markdown fences if any
                if raw.startswith("```json") or raw.startswith("```"):
                    raw = raw.strip("`\n ")
                # Attempt parse; if fails, salvage JSON substring
                try:
                    parsed = json.loads(raw)
                except json.JSONDecodeError:
                    start = raw.find("{")
                    end = raw.rfind("}")
                    if start != -1 and end != -1 and end > start:
                        parsed = json.loads(raw[start : end + 1])
                    else:
                        raise
                return Extraction(**parsed)
            except Exception:
                if attempt == 1:
                    return None
                continue
        return None


# -----------------------------
# LinkedIn automation
# -----------------------------

LOGIN_URL = "https://www.linkedin.com/login"
FEED_URL = "https://www.linkedin.com/feed/"
SEARCH_URL_TMPL = (
    "https://www.linkedin.com/search/results/content/?keywords={query}&origin=GLOBAL_SEARCH_HEADER"
)


def launch_browser(playwright, headless: bool = True) -> Tuple[Browser, BrowserContext]:
    browser = playwright.chromium.launch(headless=headless)
    context = browser.new_context(
        viewport={"width": 1366, "height": 900},
        user_agent=(
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        ),
    )
    return browser, context


def save_storage_state(context: BrowserContext, storage_state_path: str) -> None:
    try:
        context.storage_state(path=storage_state_path)
    except Exception:
        pass


def try_login_with_credentials(page: Page, email: str, password: str) -> bool:
    try:
        page.goto(LOGIN_URL, wait_until="domcontentloaded", timeout=30000)
        sleep_random(0.5, 1.2)
        page.fill("#username", email, timeout=15000)
        sleep_random(0.1, 0.4)
        page.fill("#password", password, timeout=15000)
        sleep_random(0.1, 0.4)
        page.click("button[type='submit']", timeout=15000)
        # Wait for feed or checkpoint
        try:
            page.wait_for_url("**/feed/**", timeout=30000)
            return True
        except PlaywrightTimeoutError:
            # Could be 2FA or checkpoint; give some extra time
            try:
                page.wait_for_url("**/checkpoint/**", timeout=20000)
            except PlaywrightTimeoutError:
                pass
            # Try checking nav presence as a signal of success
            try:
                page.wait_for_selector("nav", timeout=15000)
                return True
            except PlaywrightTimeoutError:
                return False
    except Exception:
        return False


def ensure_logged_in(context: BrowserContext, storage_state_path: str, email: Optional[str], password: Optional[str], headless: bool) -> bool:
    page = context.new_page()
    try:
        # If we already have cookies, try feed directly
        page.goto(FEED_URL, wait_until="domcontentloaded", timeout=30000)
        # If not redirected to login, assume logged in
        if "login" not in page.url:
            return True
    except PlaywrightTimeoutError:
        pass

    # Try credentials if provided
    if email and password:
        ok = try_login_with_credentials(page, email, password)
        if ok:
            return True

    # If headful, user may complete 2FA manually. Provide grace window.
    if not headless:
        try:
            page.wait_for_url("**/feed/**", timeout=120000)
            return True
        except PlaywrightTimeoutError:
            return False

    return False


def safe_text(el) -> Optional[str]:
    try:
        txt = el.inner_text(timeout=2000)
        return normalize_whitespace(txt)
    except Exception:
        return None


def safe_attr(el, name: str) -> Optional[str]:
    try:
        val = el.get_attribute(name, timeout=2000)
        return val
    except Exception:
        return None


def first_text(page_or_el, selector: str) -> Optional[str]:
    try:
        el = page_or_el.locator(selector).first
        if el.count() == 0:
            return None
        return safe_text(el)
    except Exception:
        return None


def first_href(page_or_el, selector: str) -> Optional[str]:
    try:
        el = page_or_el.locator(selector).first
        if el.count() == 0:
            return None
        href = safe_attr(el, "href")
        return ensure_full_url(href)
    except Exception:
        return None


def extract_post_fields(post_el) -> Dict[str, Optional[str]]:
    # Attempt multiple selectors for robustness
    post_text = (
        first_text(post_el, "div[dir='ltr']")
        or first_text(post_el, "span[dir='ltr']")
        or first_text(post_el, "p")
        or safe_text(post_el)
    )

    poster_link = (
        first_href(post_el, "a[href*='/in/']")
        or first_href(post_el, "a[href*='linkedin.com/in/']")
    )
    poster_name = (
        first_text(post_el, "a[href*='/in/']")
        or first_text(post_el, "span.feed-shared-actor__name")
        or None
    )

    # Post permalink candidates
    post_url = (
        first_href(post_el, "a[href*='/posts/']")
        or first_href(post_el, "a[href*='/activity/']")
        or first_href(post_el, "a[href*='/feed/update/urn:']")
    )

    timestamp_text = (
        first_text(post_el, "time")
        or first_text(post_el, "span.update-components-actor__sub-description")
        or None
    )

    return {
        "post_text": normalize_whitespace(post_text) if post_text else None,
        "poster_name": normalize_whitespace(poster_name) if poster_name else None,
        "poster_profile_url": ensure_full_url(poster_link) if poster_link else None,
        "post_url": ensure_full_url(post_url) if post_url else None,
        "timestamp_text": normalize_whitespace(timestamp_text) if timestamp_text else None,
    }


def scroll_and_collect_posts(page: Page, max_posts: int) -> List[Any]:
    collected = []
    seen_hashes = set()

    last_height = 0
    stagnant_scrolls = 0

    while len(collected) < max_posts and stagnant_scrolls < 5:
        try:
            page.wait_for_selector("article", timeout=8000)
        except PlaywrightTimeoutError:
            pass

        articles = page.locator("article")
        count = articles.count()
        for i in range(count):
            if len(collected) >= max_posts:
                break
            el = articles.nth(i)
            data = extract_post_fields(el)
            summary_blob = json.dumps(data, sort_keys=True)
            h = text_hash(summary_blob)
            if h in seen_hashes:
                continue
            seen_hashes.add(h)
            collected.append(data)

        # Scroll
        page.mouse.wheel(0, 2000)
        sleep_random(0.8, 1.6)

        # Heuristic to detect end
        try:
            new_height = page.evaluate("() => document.body.scrollHeight")
            if new_height == last_height:
                stagnant_scrolls += 1
            else:
                stagnant_scrolls = 0
                last_height = new_height
        except Exception:
            pass

    return collected


def search_posts(page: Page, query: str, max_posts: int) -> List[Dict[str, Optional[str]]]:
    url = SEARCH_URL_TMPL.format(query=requests.utils.quote(query))
    page.goto(url, wait_until="domcontentloaded", timeout=45000)
    sleep_random(1.2, 2.2)

    # Explicitly click the Posts tab if visible
    try:
        posts_tab = page.get_by_role("link", name="Posts")
        if posts_tab and posts_tab.count() > 0:
            posts_tab.first.click(timeout=5000)
            sleep_random(1.0, 1.8)
    except Exception:
        pass

    results = scroll_and_collect_posts(page, max_posts=max_posts)
    return results


# -----------------------------
# Excel I/O
# -----------------------------

EXPECTED_COLUMNS = [
    "timestamp",
    "post_url",
    "poster_name",
    "poster_profile_url",
    "poster_linkedin_id",
    "role_title",
    "min_years_experience",
    "max_years_experience",
    "skills",
    "location",
    "job_type",
    "contact",
    "post_excerpt",
]


def load_existing_excel(path: str) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    if not os.path.exists(path):
        return rows
    try:
        wb = load_workbook(filename=path)
        ws = wb.active
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        header = [str(h) if h is not None else "" for h in header]
        for r in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {header[i]: (r[i] if i < len(r) else None) for i in range(len(header))}
            # Normalize to expected columns
            normalized = {col: row_dict.get(col, None) for col in EXPECTED_COLUMNS}
            rows.append(normalized)
        return rows
    except Exception:
        return []


def append_and_save_excel(path: str, new_rows: List[PostRecord]) -> None:
    existing_rows = load_existing_excel(path)
    incoming_rows = [r.to_row() for r in new_rows]

    combined: List[Dict[str, Any]] = []
    seen: set[str] = set()

    def make_key(row: Dict[str, Any]) -> str:
        return f"{row.get('post_url') or ''}|{row.get('post_excerpt') or ''}"

    for row in existing_rows + incoming_rows:
        # Ensure all expected columns present
        for col in EXPECTED_COLUMNS:
            row.setdefault(col, "")
        key = make_key(row)
        if key in seen:
            continue
        seen.add(key)
        combined.append(row)

    wb = Workbook()
    ws = wb.active
    ws.append(EXPECTED_COLUMNS)
    for row in combined:
        ws.append([row.get(col, "") for col in EXPECTED_COLUMNS])
    wb.save(path)


# -----------------------------
# Main orchestration
# -----------------------------

def build_post_record(raw: Dict[str, Optional[str]], extraction: Extraction) -> PostRecord:
    poster_profile_url = raw.get("poster_profile_url")
    poster_id = extract_profile_id(poster_profile_url)

    excerpt = raw.get("post_text")
    if excerpt and len(excerpt) > 500:
        excerpt = excerpt[:497] + "..."

    return PostRecord(
        timestamp_text=raw.get("timestamp_text"),
        post_url=raw.get("post_url"),
        poster_name=raw.get("poster_name"),
        poster_profile_url=poster_profile_url,
        poster_linkedin_id=poster_id,
        role_title=extraction.role_title,
        min_years_experience=extraction.min_years_experience,
        max_years_experience=extraction.max_years_experience,
        skills=extraction.skills or [],
        location=extraction.location,
        job_type=extraction.job_type,
        contact=extraction.contact,
        post_excerpt=excerpt,
    )


def run_scrape(query: str, max_posts: int, output_excel: str, storage_state_path: str, headless: bool, email: Optional[str], password: Optional[str], ollama_url: str, ollama_model: str) -> None:
    ollama = OllamaClient(base_url=ollama_url, model=ollama_model)

    with sync_playwright() as p:
        browser, context = launch_browser(p, headless=headless)

        # Load storage state if exists
        if os.path.exists(storage_state_path):
            try:
                context.close()
                browser.close()
            except Exception:
                pass
            browser = p.chromium.launch(headless=headless)
            context = browser.new_context(
                viewport={"width": 1366, "height": 900},
                storage_state=storage_state_path,
            )

        # Ensure logged in
        ok = ensure_logged_in(context, storage_state_path, email, password, headless=headless)
        if not ok:
            print("[ERROR] Not logged in. Provide credentials via --email/--password or create a storage state via --login.", file=sys.stderr)
            try:
                context.close(); browser.close()
            except Exception:
                pass
            sys.exit(1)

        # Save updated cookies
        save_storage_state(context, storage_state_path)

        page = context.new_page()
        results = search_posts(page, query=query, max_posts=max_posts)

        print(f"[INFO] Collected {len(results)} raw posts; sending to LLM for filtering/extraction...")

        kept: List[PostRecord] = []
        for idx, raw in enumerate(results, start=1):
            text = raw.get("post_text") or ""
            if not text.strip():
                continue

            extraction = ollama.extract(text)
            if extraction is None:
                # Fallback heuristic: basic keyword check
                t = text.lower()
                has_role = ("data analyst" in t) or ("junior data analyst" in t)
                fresher_markers = ["0-2", "0 to 2", "freshers", "fresher", "entry level", "junior"]
                has_exp = any(m in t for m in fresher_markers)
                is_relevant = has_role and has_exp
                extraction = Extraction(
                    role_title="Data Analyst" if has_role else None,
                    min_years_experience=0 if has_exp else None,
                    max_years_experience=2 if has_exp else None,
                    skills=[],
                    location=None,
                    job_type=None,
                    contact=None,
                    verdict_relevant=is_relevant,
                )

            if not extraction.verdict_relevant:
                continue

            record = build_post_record(raw, extraction)
            kept.append(record)

            if idx % 5 == 0:
                sleep_random(0.8, 1.5)

        if kept:
            append_and_save_excel(output_excel, kept)
            print(f"[OK] Wrote {len(kept)} records to {output_excel}")
        else:
            print("[INFO] No relevant posts found based on the criteria.")

        try:
            context.close(); browser.close()
        except Exception:
            pass


def run_login_only(storage_state_path: str, email: Optional[str], password: Optional[str], headless: bool) -> None:
    with sync_playwright() as p:
        browser, context = launch_browser(p, headless=headless)
        page = context.new_page()

        success = False
        if email and password:
            success = try_login_with_credentials(page, email, password)

        if not success:
            # If headless, cannot do manual; bail
            if headless:
                print("[ERROR] Headless login failed and manual login not possible in headless mode.", file=sys.stderr)
                try:
                    context.close(); browser.close()
                except Exception:
                    pass
                sys.exit(1)
            # Allow manual login window
            print("[ACTION] Please complete login in the opened browser window. Waiting up to 2 minutes...")
            page.goto(LOGIN_URL, wait_until="domcontentloaded", timeout=30000)
            try:
                page.wait_for_url("**/feed/**", timeout=120000)
                success = True
            except PlaywrightTimeoutError:
                success = False

        if success:
            save_storage_state(context, storage_state_path)
            print(f"[OK] Saved login session to {storage_state_path}")
        else:
            print("[ERROR] Login not completed.", file=sys.stderr)
            sys.exit(1)

        try:
            context.close(); browser.close()
        except Exception:
            pass


# -----------------------------
# CLI
# -----------------------------

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="LinkedIn posts scraper + local LLM extractor (Data Analyst 0-2 yrs)")
    parser.add_argument("--query", default="Data Analyst hiring", help="Search query text for posts")
    parser.add_argument("--max-posts", type=int, default=40, help="Maximum number of posts to scan")
    parser.add_argument("--output", default="linkedin_data_analyst_posts.xlsx", help="Output Excel path")
    parser.add_argument("--storage-state", default="storage_state.json", help="Path to persist LinkedIn cookies/session")

    headless_group = parser.add_mutually_exclusive_group()
    headless_group.add_argument("--headless", dest="headless", action="store_true", help="Run browser headless (default)")
    headless_group.add_argument("--no-headless", dest="headless", action="store_false", help="Run with visible browser window")
    parser.set_defaults(headless=True)

    parser.add_argument("--login", action="store_true", help="Perform login flow and save storage state, then exit")
    parser.add_argument("--email", default=os.environ.get("LINKEDIN_EMAIL"), help="LinkedIn email (or set LINKEDIN_EMAIL env)")
    parser.add_argument("--password", default=os.environ.get("LINKEDIN_PASSWORD"), help="LinkedIn password (or set LINKEDIN_PASSWORD env)")

    parser.add_argument("--ollama-url", default=os.environ.get("OLLAMA_HOST", "http://localhost:11434"), help="Ollama base URL")
    parser.add_argument("--ollama-model", default=os.environ.get("OLLAMA_MODEL", "llama3"), help="Ollama model name (e.g., llama3, mistral)")

    return parser.parse_args()


def main() -> None:
    args = parse_args()

    if args.login:
        run_login_only(
            storage_state_path=args.storage_state,
            email=args.email,
            password=args.password,
            headless=args.headless,
        )
        return

    run_scrape(
        query=args.query,
        max_posts=args.max_posts,
        output_excel=args.output,
        storage_state_path=args.storage_state,
        headless=args.headless,
        email=args.email,
        password=args.password,
        ollama_url=args.ollama_url,
        ollama_model=args.ollama_model,
    )


if __name__ == "__main__":
    main()